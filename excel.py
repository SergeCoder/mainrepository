import openpyxl as op
import pandas as pd
import os
import re
from datetime import datetime

unclear_list = []
clear_list = []

clear_number_list = []
log_list = []
check_list = []

other_data_list = []
clear_other_list = []

copy_data_list = []
copy_other_list = []
clear_copy_list = []
unfound_numbers = []

check_files = []
report_files = []
xlsx_files = []
xml_files = []

check_dates = []
report_dates = []
xlsx_dates = []
xml_dates = []

def find_report():
    """открытие отчета"""
    all_files = os.listdir()
    for file in all_files:
        date = re.findall('([0-9]{4})', file)
        try:
            if 'отчет' in file.lower():  # поиск xlsx отчета
                report_files.append(file)
                report_dates.append(file[-9:-5:])
        except IndexError:
            continue

def find_books(xlsx_files, xml_files, xlsx_dates, xml_dates):
    """открытие книг"""
    all_files = os.listdir()
    for file in all_files:
        date = re.findall('([0-9]{2}\\.[0-9]{2}\\.[0-9]{4})', file)
        try:
            if ('бронь' in str(file).lower()) and (date[0] in str(file).lower()): # поиск xlsx файла
                xlsx_files.append(file)
                xlsx_dates.append(file[-15:-5:])
        except IndexError:
            continue

        if '.xml' in file: # поиск xml файла
            xml_files.append(file)
            xml_dates.append(file[-15:-4:])

find_books(xlsx_files=xlsx_files, xml_files=xml_files, xlsx_dates=xlsx_dates, xml_dates=xml_dates)
find_report()

xlsx_file = (f'{os.path.dirname(os.path.abspath(__file__))}\\'
             f'{xlsx_files[xlsx_dates.index(max(xlsx_dates))]}')

xml_file = (f'{os.path.dirname(os.path.abspath(__file__))}\\'
            f'{xml_files[xml_dates.index(max(xml_dates))]}')

report_file = (f'{os.path.dirname(os.path.abspath(__file__))}\\'
               f'{report_files[report_dates.index(max(report_dates))]}')

date = str(datetime.date(datetime.now())).split('-')
new_file = f'Бронь {date[2]}.{date[1]}.{date[0]}.xlsx'
unfound_file = f'Проверка {date[2]}.{date[1]}.{date[0]}.xlsx'
log_file = f'Log {date[2]}.{date[1]}.{date[0]}.xlsx'

def collect_data(schet_list, other_data, clear_other_data, file, sheet):
    """сбор данных из прайс-листа"""
    info = pd.read_excel(file, sheet) # поиск количества строк
    count = info.count().to_list()
    count = count[6]

    df = pd.read_excel(file, sheet_name=sheet)
    numbers = df.iloc[:, 6] # поиск информации в таблице
    names = df.iloc[:, 1]
    inns = df.iloc[:, 2]
    kbs = df.iloc[:, 3]
    start_dates = df.iloc[:, 4]
    end_dates = df.iloc[:, 5]
    schets = df.iloc[:, 6]
    prices_schets = df.iloc[:, 7]
    discounts = df.iloc[:, 8]
    makets = df.iloc[:, 9]
    prices_makets = df.iloc[:, 10]

    for number in range(2, count+2): # форматирование счетов и данных
        schetnomer = numbers[number]

        try:
            schetnomer = schetnomer.split(' ')[1]
        except IndexError:
            continue

        name = names[number]
        inn = inns[number]
        kb = kbs[number]
        start_date = start_dates[number]
        end_date = end_dates[number]
        schet = schets[number]
        price_schet = prices_schets[number]
        discount = discounts[number]
        maket = makets[number]
        price_maket = prices_makets[number]

        if schetnomer != '':
            schet_list.append(schetnomer)

        if not [name, inn, kb, start_date, end_date, schet,
                price_schet, discount, maket, price_maket] in other_data:
            other_data.append([name, inn, kb, start_date, end_date, schet,
                               price_schet, discount, maket, price_maket])

    for data in other_data:
        for clear_schet in schet_list:
            if clear_schet in data[5]:
                clear_other_data.append([data[0], data[1], data[2], data[3], data[4],
                                         data[5], data[6], data[7], data[8], data[9]])

def create_table(file):
    """создание новой таблицы"""
    book = op.Workbook()
    sheet = book.active

    sheet.delete_cols(0, 25)

    book.save(file)

    book.save(file)
    book.close()

def write_data(file, number_list, other_data_list):
    """запись информации в новую книгу"""
    book = op.load_workbook(file)
    sheet = book[pd.ExcelFile(file).sheet_names[0]]

    sheet['A4'] = 'Телефон'
    sheet['B4'] = 'Контрагент'
    sheet['C4'] = 'ИНН'
    sheet['D4'] = 'Тип брони'
    sheet['E4'] = 'Дата начала'
    sheet['F4'] = 'Дата конца'
    sheet['G4'] = 'Счет'
    sheet['H4'] = 'Сумма счета'
    sheet['I4'] = 'Средняя скидка'
    sheet['J4'] = 'Макет'
    sheet['K4'] = 'Сумма макета'

    row = 5
    for number in number_list:
        sheet[f'A{row}'] = int(number)
        row += 1

    row = 5
    for data in other_data_list:
        sheet[f'B{row}'] = data[0]
        sheet[f'C{row}'] = data[1]
        sheet[f'D{row}'] = data[2]
        sheet[f'E{row}'] = data[3]
        sheet[f'F{row}'] = data[4]
        sheet[f'G{row}'] = data[5]
        sheet[f'H{row}'] = data[6]
        sheet[f'I{row}'] = data[7]
        sheet[f'J{row}'] = data[8]
        sheet[f'K{row}'] = data[9]

        row += 1

    book.save(file)
    book.close()

def collect_unfound(copy_list, copy_other_data, clear_data,
                    unfound_numbers, clear_copy_data, file, sheet):
    """ищет счета в отчете"""
    info = pd.read_excel(file, sheet)  # поиск количества строк
    count = info.count().to_list()
    count = count[0]

    df = pd.read_excel(file, sheet_name=sheet)
    numbers = df.iloc[:, 2]  # поиск информации в таблице
    names = df.iloc[:, 1]
    strings = df.iloc[:, 0]

    for number in range(2, count + 2):  # форматирование счетов и данных
        schetnomer = numbers[number]
        name = names[number]
        string = strings[number]

        try:
            if string.isdigit():
                copy_list.append(schetnomer)
        except IndexError:
            continue

        if (not [string, name] in copy_other_data) and string.isdigit():
            copy_other_data.append([string, name])

    for data in copy_list:
        clear_copy_data.append(data.split(' ')[2])

    for clear_schet in clear_copy_data:
        if (not (clear_schet in clear_data) and
            not (clear_schet in unfound_numbers)):
            unfound_numbers.append(clear_schet)

    print(unfound_numbers)
    print(copy_other_list)

def write_unfound(file, unfound_numbers, copy_other_list):
    """запись информации в файл для не найденных номеров"""
    book = op.load_workbook(file)
    sheet = book[pd.ExcelFile(file).sheet_names[0]]

    sheet['B2'] = 'Не найденные данные в Брони'

    sheet['A4'] = '№ п/п'
    sheet['B4'] = 'Контрагент'
    sheet['C4'] = '№ счета'

    row = 5
    for number in unfound_numbers:
        sheet[f'C{row}'] = number
        sheet[f'A{row}'] = copy_other_list[unfound_numbers.index(number)][0]
        sheet[f'B{row}'] = copy_other_list[unfound_numbers.index(number)][1]

        row += 1

    book.save(file)
    book.close()

def check_numbers(unclear_data, log_list, check_list):
    """проверяет номера из xlsx"""
    for number in unclear_data:
        if not number in log_list:
            if not number in check_list:
                check_list.append(number)

def write_log(file, check_list):
    """запись информации в файл для не найденных номеров"""
    book = op.load_workbook(file)
    sheet = book[pd.ExcelFile(file).sheet_names[0]]

    sheet['A2'] = 'Не найденные данные в xml'

    sheet['A4'] = '№ счета'

    row = 5
    for number in check_list:
        sheet[f'A{row}'] = number

        row += 1

    book.save(file)
    book.close()